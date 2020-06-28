from openpyxl import load_workbook

wb = load_workbook('record.xlsx')
print(wb.sheetnames)


work_sheet = wb['math']
print(work_sheet.cell(2,2).value)
for column_index, column in enumerate(work_sheet.columns):
    total = 0
    for row_item in column:
        total += row_item.value

    size = len(column)
    average = total / size
    work_sheet.cell(size+1, column_index+1).value = average

wb.save('record.xlsx')
